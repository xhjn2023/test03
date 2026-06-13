import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy
import os
import sys
import re
from datetime import datetime

TARGET_SHEET_KEYWORDS = ['寿险2021智慧银保行销支持系统', '寿险2024版银保品质管理系统']
HEADER_KEYWORDS = ['用例编号', '测试项']
UAT_FIXED_TESTER = ''
UAT_FIXED_ROLE = '用户'
UAT_DEFAULT_RESULT = '是'
CONFORMITY_FIXED_TESTER = '王嘉裕'
CONFORMITY_DEMAND_NAME = '适当性优化'
FONT_SONGTI_14 = Font(name='宋体', size=14, bold=True, color='FFFFFF')
FONT_SONGTI_12 = Font(name='宋体', size=12, bold=True, color='FFFFFF')
FONT_SONGTI_12_BODY = Font(name='宋体', size=12)
FONT_SONGTI_14_HEADER = Font(name='宋体', size=14, bold=True, color='FFFFFF')
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
HEADER_FILL_UAT = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type=None)
FILL_CONFORMITY_TITLE = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
FILL_CONFORMITY_HEADER = PatternFill(start_color='8DB4E2', end_color='8DB4E2', fill_type='solid')
FILL_ISSUE_TITLE = PatternFill(start_color='F57C1B', end_color='F57C1B', fill_type='solid')
FILL_ISSUE_HEADER = PatternFill(start_color='F8A25C', end_color='F8A25C', fill_type='solid')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)


def resolve_source_sheet(wb):
    for sn in wb.sheetnames:
        for kw in TARGET_SHEET_KEYWORDS:
            if kw in sn:
                return sn
    for sn in wb.sheetnames:
        for kw in HEADER_KEYWORDS:
            ws = wb[sn]
            for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=True):
                row_str = '|'.join(str(v or '') for v in row)
                if all(kw in row_str for kw in HEADER_KEYWORDS):
                    return sn
    return wb.sheetnames[2]


def find_header_row(ws):
    for row_idx in range(1, min(20, ws.max_row) + 1):
        row_vals = [str(ws.cell(row=row_idx, column=c).value or '') for c in range(1, ws.max_column + 1)]
        row_str = '|'.join(row_vals)
        if all(kw in row_str for kw in HEADER_KEYWORDS):
            return row_idx
    return None


def col_map_from_header(ws, header_row):
    mapping = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        if val:
            mapping[val.strip()] = c
    return mapping


def get_file_prefix(filepath):
    basename = os.path.splitext(os.path.basename(filepath))[0]
    m = re.match(r'^(.+?)单元测试报告', basename)
    if m:
        return m.group(1)
    return basename


def read_data_rows(ws, header_row, col_map):
    rows = []
    key_col_candidates = ['用例编号']
    key_col = None
    for kc in key_col_candidates:
        if kc in col_map:
            key_col = col_map[kc]
            break

    for r in range(header_row + 1, ws.max_row + 1):
        if key_col:
            key_val = ws.cell(row=r, column=key_col).value
            if key_val is None or str(key_val).strip() == '':
                continue
        row_data = {}
        for name, col_idx in col_map.items():
            row_data[name] = ws.cell(row=r, column=col_idx).value
        rows.append(row_data)
    return rows


def safe_str(val):
    if val is None:
        return ''
    return str(val).strip()


def create_uat_excel(data_rows, col_map, prefix, output_dir, source_filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'UAT测试用例'

    headers = ['功能点*', '测试案例编号*', '测试案例名称*', '系统角色*',
               '测试步骤及步骤描述*', '预期结果*', '测试结果（是否通过）',
               '问题描述（如未通过）', '测试负责人']

    header_font = Font(name='宋体', size=12, bold=True)
    body_font = Font(name='宋体', size=12)

    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.font = header_font
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    current_row = 2
    for row_data in data_rows:
        case_id = safe_str(row_data.get('用例编号', ''))
        test_item = safe_str(row_data.get('测试项', ''))
        case_desc = safe_str(row_data.get('用例描述', ''))
        test_steps = safe_str(row_data.get('测试步骤', ''))
        expected_result = safe_str(row_data.get('预期结果', ''))

        uat_row = [
            test_item,
            case_id,
            case_desc,
            UAT_FIXED_ROLE,
            test_steps,
            expected_result,
            UAT_DEFAULT_RESULT,
            '',
            UAT_FIXED_TESTER
        ]

        for c_idx, val in enumerate(uat_row, 1):
            cell = ws.cell(row=current_row, column=c_idx, value=val)
            cell.font = body_font
            cell.alignment = LEFT_ALIGN
            cell.border = THIN_BORDER

        current_row += 1

    col_widths = [20, 15, 40, 10, 60, 30, 15, 25, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output_path = os.path.join(output_dir, f'UAT测试用例-{prefix}.xlsx')
    wb.save(output_path)
    return output_path


def create_conformity_excel(data_rows, col_map, prefix, output_dir, source_filename):
    wb = openpyxl.Workbook()

    # --- Sheet 1: 测试情况 ---
    ws1 = wb.active
    ws1.title = '测试情况'

    ws1.merge_cells('A1:I1')
    title_cell = ws1.cell(row=1, column=1, value='需求符合度测试')
    title_cell.font = FONT_SONGTI_14
    title_cell.fill = FILL_CONFORMITY_TITLE
    title_cell.alignment = CENTER_ALIGN
    ws1.row_dimensions[1].height = 44.2

    conf_headers = ['需求号', '需求名称', '序号', '功能点简述',
                    '主要路径', '测试结果', '测试时间', '测试人', '工作量']
    for c_idx, h in enumerate(conf_headers, 1):
        cell = ws1.cell(row=2, column=c_idx, value=h)
        cell.font = Font(name='宋体', size=14, bold=True, color='FFFFFF')
        cell.fill = FILL_CONFORMITY_HEADER
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    ws1.row_dimensions[2].height = 42

    today_str = datetime.now().strftime('%m月%d号')

    current_row = 3
    for row_data in data_rows:
        case_id = safe_str(row_data.get('用例编号', ''))
        test_item = safe_str(row_data.get('测试项', ''))
        test_steps = safe_str(row_data.get('测试步骤', ''))
        exec_status = safe_str(row_data.get('执行状态', ''))

        if exec_status in ('通过', 'Pass'):
            test_result = '成功'
        else:
            test_result = exec_status

        conf_row = [
            '',
            prefix,
            case_id,
            test_item,
            test_steps,
            test_result,
            today_str,
            CONFORMITY_FIXED_TESTER,
            ''
        ]

        for c_idx, val in enumerate(conf_row, 1):
            cell = ws1.cell(row=current_row, column=c_idx, value=val)
            cell.font = FONT_SONGTI_12_BODY
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        current_row += 1

    col_widths_1 = [30.44, 27.78, 10.78, 42.44, 60.44, 22.22, 13.22, 9.89, 21.11]
    for i, w in enumerate(col_widths_1, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # --- Sheet 2: 测试问题归纳 ---
    ws2 = wb.create_sheet(title='测试问题归纳')

    ws2.merge_cells('A1:D1')
    issue_title_cell = ws2.cell(row=1, column=1, value='测试问题归纳')
    issue_title_cell.font = FONT_SONGTI_12
    issue_title_cell.fill = FILL_ISSUE_TITLE
    issue_title_cell.alignment = CENTER_ALIGN
    ws2.row_dimensions[1].height = 21.95

    issue_headers = ['序号', '问题描述', '问题类型', '备注']
    for c_idx, h in enumerate(issue_headers, 1):
        cell = ws2.cell(row=2, column=c_idx, value=h)
        cell.font = Font(name='宋体', size=12, bold=True, color='FFFFFF')
        cell.fill = FILL_ISSUE_HEADER
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    ws2.row_dimensions[2].height = 21.95

    current_row = 3
    for row_data in data_rows:
        exec_status = safe_str(row_data.get('执行状态', ''))
        if exec_status not in ('通过', 'Pass', '成功', ''):
            case_id = safe_str(row_data.get('用例编号', ''))
            test_item = safe_str(row_data.get('测试项', ''))
            ws2.cell(row=current_row, column=1, value=case_id).border = THIN_BORDER
            ws2.cell(row=current_row, column=1).font = FONT_SONGTI_12_BODY
            ws2.cell(row=current_row, column=1).alignment = CENTER_ALIGN
            ws2.cell(row=current_row, column=2, value=f'{test_item} - 执行状态: {exec_status}').border = THIN_BORDER
            ws2.cell(row=current_row, column=2).font = FONT_SONGTI_12_BODY
            ws2.cell(row=current_row, column=2).alignment = LEFT_ALIGN
            ws2.cell(row=current_row, column=3).border = THIN_BORDER
            ws2.cell(row=current_row, column=4).border = THIN_BORDER
            current_row += 1

    col_widths_2 = [10, 50, 20, 20]
    for i, w in enumerate(col_widths_2, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output_path = os.path.join(output_dir, f'符合度测试用例-{prefix}.xlsx')
    wb.save(output_path)
    return output_path


def main():
    if len(sys.argv) > 1:
        file_paths = [sys.argv[1]]
    else:
        cwd = os.getcwd()
        file_paths = [
            os.path.join(cwd, f)
            for f in os.listdir(cwd)
            if '单元测试报告' in f and f.endswith('.xlsx')
        ]

    if not file_paths:
        print('未找到单元测试报告文件')
        sys.exit(1)

    for filepath in file_paths:
        if not os.path.isfile(filepath):
            print(f'文件不存在: {filepath}')
            continue

        filename = os.path.basename(filepath)
        output_dir = os.path.dirname(filepath)
        prefix = get_file_prefix(filepath)
        print(f'\n处理文件: {filename}')
        print(f'前缀: {prefix}')

        wb = openpyxl.load_workbook(filepath, data_only=True)

        sheet_name = resolve_source_sheet(wb)
        ws = wb[sheet_name]
        print(f'使用Sheet: {sheet_name}')

        header_row = find_header_row(ws)
        if not header_row:
            print(f'错误: 未找到表头行')
            continue
        print(f'表头行: {header_row}')

        cm = col_map_from_header(ws, header_row)
        print(f'列映射: {cm}')

        data_rows = read_data_rows(ws, header_row, cm)
        print(f'数据行数: {len(data_rows)}')

        uat_path = create_uat_excel(data_rows, cm, prefix, output_dir, filename)
        print(f'UAT用例已生成: {uat_path}')

        conf_path = create_conformity_excel(data_rows, cm, prefix, output_dir, filename)
        print(f'符合度用例已生成: {conf_path}')

    print('\n转换完成!')


if __name__ == '__main__':
    main()
